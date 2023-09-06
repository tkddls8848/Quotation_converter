import xml.etree.ElementTree as ETree
import xmltodict
from openpyxl import Workbook

tree = ETree.parse("C://Users/tkddl/Desktop/test.xml")
root = tree.getroot()

tmp = []
for ele in root:
    if ele.tag == "CFData":
        tmp = list(ele)
tmp_row = 1
write_wb = Workbook()
write_ws = write_wb.create_sheet("TEST")
write_ws = write_wb.active

for data in tmp:
    if data.tag == "ProductLineItem":
        top_data_product_details = dict(xmltodict.parse(ETree.tostring(data)))["ProductLineItem"]["ProductIdentification"]["PartnerProductIdentification"]
        top_data_product_price = dict(xmltodict.parse(ETree.tostring(data)))["ProductLineItem"]["UnitListPrice"]["FinancialAmount"]["MonetaryAmount"].replace(",", "")
        top_data_product_quantity = dict(xmltodict.parse(ETree.tostring(data)))["ProductLineItem"]["Quantity"]
        arr = [top_data_product_details["ProductTypeCode"], top_data_product_details["ProprietaryProductIdentifier"], top_data_product_details["ProductDescription"], top_data_product_price, top_data_product_quantity]
        if top_data_product_price == "N/C":
            arr.append("N/C")
        else:
            arr.append(float(top_data_product_price)*float(top_data_product_quantity))
        tmp_col = 1
        for _ in arr:
            write_ws.cell(tmp_row, tmp_col, _)
            tmp_col += 1
        tmp_row += 1
        write_wb.save("C://Users/tkddl/Desktop/output.xlsx")

        tmp = {}
        for d in data:
            da = dict(xmltodict.parse(ETree.tostring(d)))
            if d.tag == "ProductSubLineItem":
                t = str(da)
                sub_product_quantity = eval(t)["ProductSubLineItem"]["Quantity"]
                sub_product_info_details = eval(t)["ProductSubLineItem"]["ProductIdentification"]["PartnerProductIdentification"]
                sub_product_price_details = eval(t)["ProductSubLineItem"]["UnitListPrice"]["FinancialAmount"]["MonetaryAmount"].replace(",", "")

                arr = [sub_product_info_details["ProductTypeCode"], sub_product_info_details["ProprietaryProductIdentifier"], sub_product_info_details["ProductDescription"], sub_product_price_details, sub_product_quantity]
                if sub_product_price_details == "N/C":
                    arr.append("N/C")
                else:
                    arr.append(float(sub_product_price_details) * float(sub_product_quantity))
                tmp_col = 1
                for _ in arr:
                    write_ws.cell(tmp_row, tmp_col, _)
                    tmp_col += 1
                tmp_row += 1

                write_wb.save("C://Users/tkddl/Desktop/output.xlsx")
        write_ws.cell(tmp_row, 1, "TOTAL")
        tmp_row += 1
